import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
from backend.clients_dao import create_cliente, read_clientes, update_cliente, delete_cliente
from backend.incidents_dao import create_incidente, read_incidentes, update_incidente, delete_incidente
from backend.installations_dao import create_instalacion, read_instalaciones, update_instalacion, delete_instalacion
from backend.personal_dao import create_personal, read_personal, update_personal, delete_personal
from backend.db import init_pool
import openpyxl




class SecureSysApp(tk.Tk):

    def __init__(self):
        super().__init__()
        self.title('SecureSys - Conectado')
        self.geometry('1000x650')

        # init DB pool (lee variables de entorno)
        try:
            init_pool()
        except Exception as e:
            print('Advertencia: No se pudo inicializar pool de conexiones:', e)

        # menú
        menubar = tk.Menu(self)
        self.config(menu=menubar)
        mod_clientes = tk.Menu(menubar, tearoff=0)
        mod_incidentes = tk.Menu(menubar, tearoff=0)
        mod_personal = tk.Menu(menubar, tearoff=0)
        mod_instalaciones = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label='Clientes', menu=mod_clientes)
        menubar.add_cascade(label='Incidentes', menu=mod_incidentes)
        menubar.add_cascade(label='Instalaciones', menu=mod_instalaciones)
        menubar.add_cascade(label='Personal Seguridad', menu=mod_personal)
        mod_clientes.add_command(label='Administrar Clientes', command=self.show_clientes_view)
        mod_incidentes.add_command(label='Administrar Incidentes', command=self.show_incidentes_view)
        mod_instalaciones.add_command(label='Administrar Instalaciones', command=self.show_instalaciones)
        mod_personal.add_command(label='Administrar Personal', command=self.show_personal_view)

        # sidebar
        self.sidebar = tk.Frame(self, bg='#2c3e50', width=220)
        self.sidebar.pack(fill='y', side='left')
        lbl = tk.Label(self.sidebar, text='Operaciones', bg='#2c3e50', fg='white', font=('Arial',12,'bold'))
        lbl.pack(pady=10)
        tk.Button(self.sidebar, text='Clientes', width=20, command=self.show_clientes_view).pack(pady=5)
        tk.Button(self.sidebar, text='Incidentes', width=20, command=self.show_incidentes_view).pack(pady=5)
        tk.Button(self.sidebar, text='Instalaciones', width=20, command=self.show_instalaciones).pack(pady=5)
        tk.Button(self.sidebar, text='Personal', width=20, command=self.show_personal_view).pack(pady=5)

        # content
        self.content = tk.Frame(self, bg='#ecf0f1')
        self.content.pack(fill='both', expand=True, side='right')
        self.current_view = None
        self.show_welcome()

    def clear_content(self):
        for child in self.content.winfo_children():
            child.destroy()

    def show_welcome(self):
        self.clear_content()
        tk.Label(self.content, text='Bienvenido a SecureSys', font=('Arial',16), bg='#ecf0f1').pack(pady=20)

    # ---------------- Clientes view ----------------
    def show_clientes_view(self):
        self.clear_content()
        frame = tk.Frame(self.content, bg='#ecf0f1', padx=10, pady=10)
        frame.pack(fill='both', expand=True)

        # Formulario crear cliente
        form = tk.LabelFrame(frame, text='Crear / Actualizar Cliente', padx=10, pady=10)
        form.pack(fill='x', padx=10, pady=10)
        entries = {}
        fields = [('tipo','Tipo'),('razon_social','Razón Social'),('documento_fiscal','Documento Fiscal'),('direccion','Dirección'),('telefono','Teléfono'),('contacto_emergencia','Contacto Emergencia'),('nivel_riesgo','Nivel Riesgo')]
        for i,(k,label) in enumerate(fields):
            tk.Label(form, text=label).grid(row=i, column=0, sticky='w')
            ent = tk.Entry(form, width=50)
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[k]=ent

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=8)
        tk.Button(btn_frame, text='Crear Cliente', command=lambda: self.create_cliente(entries)).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Actualizar Cliente', command=lambda: self.update_cliente(entries)).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Eliminar Cliente', command=self.delete_cliente).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Refrescar Lista', command=self.load_clientes).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Exportar Excel', command=self.export_clientes_excel).pack(side='left', padx=5)

        # Tabla de clientes
        table_frame = tk.LabelFrame(frame, text='Clientes - Lista', padx=5, pady=5)
        table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        cols = ('cliente_id','tipo','razon_social','documento_fiscal','telefono','nivel_riesgo','fecha_inicio')
        tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120)
        tree.pack(fill='both', expand=True)
        tree.bind('<<TreeviewSelect>>', lambda e: self.on_cliente_select(e, tree, entries))

        self.client_tree = tree
        self.load_clientes()

    def create_cliente(self, entries):
        try:
            data = {k:entries[k].get().strip() for k in entries}
            if not data['documento_fiscal']:
                messagebox.showwarning('Validación','Documento fiscal es requerido.')
                return
            nid = create_cliente(data['tipo'], data['razon_social'], data['documento_fiscal'], data['direccion'], data['telefono'], data['contacto_emergencia'], data['nivel_riesgo'])
            messagebox.showinfo('Éxito', f'Cliente creado con ID {nid}')
            self.load_clientes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo crear cliente: {e}')

    def load_clientes(self):
        try:
            rows = read_clientes(200)
            self.client_tree.delete(*self.client_tree.get_children())
            for r in rows:
                self.client_tree.insert('', 'end', values=(r['cliente_id'], r['tipo'], r['razon_social'], r['documento_fiscal'], r.get('telefono',''), r.get('nivel_riesgo',''), r.get('fecha_inicio','')))
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo cargar clientes: {e}')

    def on_cliente_select(self, event, tree, entries):
        try:
            sel = tree.selection()
            if not sel: return
            vals = tree.item(sel[0], 'values')
            # map to entries
            mapping = ['cliente_id','tipo','razon_social','documento_fiscal','telefono','nivel_riesgo','fecha_inicio']
            # set entry values where applicable
            entries['tipo'].delete(0,'end'); entries['tipo'].insert(0, vals[1])
            entries['razon_social'].delete(0,'end'); entries['razon_social'].insert(0, vals[2])
            entries['documento_fiscal'].delete(0,'end'); entries['documento_fiscal'].insert(0, vals[3])
            entries['direccion'].delete(0,'end')
            entries['telefono'].delete(0,'end'); entries['telefono'].insert(0, vals[4])
            entries['contacto_emergencia'].delete(0,'end')
            entries['nivel_riesgo'].delete(0,'end'); entries['nivel_riesgo'].insert(0, vals[5])
            # store selected id
            self.selected_cliente_id = vals[0]
        except Exception as e:
            print('on_select error', e)

    def update_cliente(self, entries):
        try:
            if not hasattr(self, 'selected_cliente_id'):
                messagebox.showwarning('Seleccionar','Seleccione primero un cliente de la lista.')
                return
            data = {k:entries[k].get().strip() for k in entries}
            cnt = update_cliente(self.selected_cliente_id, tipo=data['tipo'], razon_social=data['razon_social'], documento_fiscal=data['documento_fiscal'], direccion=data['direccion'], telefono=data['telefono'], contacto_emergencia=data['contacto_emergencia'], nivel_riesgo=data['nivel_riesgo'])
            messagebox.showinfo('Éxito', f'{cnt} registro(s) actualizados.')
            self.load_clientes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo actualizar: {e}')

    def delete_cliente(self):
        try:
            if not hasattr(self, 'selected_cliente_id'):
                messagebox.showwarning('Seleccionar','Seleccione primero un cliente de la lista.')
                return
            if messagebox.askyesno('Confirmar','¿Eliminar cliente seleccionado?'):
                cnt = delete_cliente(self.selected_cliente_id)
                messagebox.showinfo('Éxito', f'{cnt} registro(s) eliminados.')
                self.load_clientes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo eliminar: {e}')
    def export_clientes_excel(self):
        try:
            rows = read_clientes(1000)
            if not rows:
                messagebox.showwarning("Atención", "No hay datos para exportar.")
                return

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar como...",
                initialfile="clientes.xlsx"
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Clientes"

            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Exportado a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")



    # ---------------- Incidentes view ----------------
    def show_incidentes_view(self):
        self.clear_content()
        frame = tk.Frame(self.content, bg='#ecf0f1', padx=10, pady=10)
        frame.pack(fill='both', expand=True)

        form = tk.LabelFrame(frame, text='Registrar Incidente', padx=10, pady=10)
        form.pack(fill='x', padx=10, pady=10)
        entries = {}
        fields = [('instalacion_id','Instalación ID'),('fecha_hora','Fecha y Hora (YYYY-MM-DD HH:MM:SS)'),('zona','Zona'),('tipo','Tipo'),('descripcion','Descripción'),('personal_involucrado','Personal Involucrado'),('respuesta_implementada','Respuesta Implementada'),('tiempo_reaccion','Tiempo Reacción (segundos)')]
        for i,(k,label) in enumerate(fields):
            tk.Label(form, text=label).grid(row=i, column=0, sticky='w')
            ent = tk.Entry(form, width=60)
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[k]=ent

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=8)
        tk.Button(btn_frame, text='Crear Incidente', command=lambda: self.create_incidente(entries)).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Actualizar Incidente', command=lambda: self.update_incidente(entries)).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Eliminar Incidente', command=self.delete_incidente).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Refrescar Lista', command=self.load_incidentes).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Exportar Excel', command=self.export_incidentes_excel).pack(side='left', padx=5)

        table_frame = tk.LabelFrame(frame, text='Incidentes - Lista', padx=5, pady=5)
        table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        cols = ('incidente_id','fecha_hora','instalacion_id','zona','tipo','tiempo_reaccion')
        tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=140)
        tree.pack(fill='both', expand=True)
        tree.bind('<<TreeviewSelect>>', lambda e: self.on_incidente_select(e, tree, entries))

        self.incident_tree = tree
        self.load_incidentes()

    def create_incidente(self, entries):
        try:
            data = {k:entries[k].get().strip() for k in entries}
            # validate date format
            try:
                datetime.strptime(data['fecha_hora'],'%Y-%m-%d %H:%M:%S')
            except Exception:
                messagebox.showwarning('Validación','Formato de fecha inválido. Use YYYY-MM-DD HH:MM:SS')
                return
            nid = create_incidente(int(data['instalacion_id']), data['fecha_hora'], data['zona'], data['tipo'], data['descripcion'], data['personal_involucrado'], data['respuesta_implementada'], int(data['tiempo_reaccion'] or 0))
            messagebox.showinfo('Éxito', f'Incidente creado con ID {nid}')
            self.load_incidentes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo crear incidente: {e}')

    def load_incidentes(self):
        try:
            rows = read_incidentes(200)
            self.incident_tree.delete(*self.incident_tree.get_children())
            for r in rows:
                self.incident_tree.insert('', 'end', values=(r['incidente_id'], r['fecha_hora'], r['instalacion_id'], r.get('zona',''), r.get('tipo',''), r.get('tiempo_reaccion',0)))
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo cargar incidentes: {e}')

    def on_incidente_select(self, event, tree, entries):
        try:
            sel = tree.selection()
            if not sel: return
            vals = tree.item(sel[0], 'values')
            entries['instalacion_id'].delete(0,'end'); entries['instalacion_id'].insert(0, vals[2])
            entries['fecha_hora'].delete(0,'end'); entries['fecha_hora'].insert(0, vals[1])
            entries['zona'].delete(0,'end'); entries['zona'].insert(0, vals[3])
            entries['tipo'].delete(0,'end'); entries['tipo'].insert(0, vals[4])
            entries['tiempo_reaccion'].delete(0,'end'); entries['tiempo_reaccion'].insert(0, vals[5])
            self.selected_incidente_id = vals[0]
        except Exception as e:
            print('on_incidente_select error', e)

    def update_incidente(self, entries):
        try:
            if not hasattr(self, 'selected_incidente_id'):
                messagebox.showwarning('Seleccionar','Seleccione primero un incidente de la lista.')
                return
            data = {k:entries[k].get().strip() for k in entries}
            cnt = update_incidente(self.selected_incidente_id, instalacion_id=int(data['instalacion_id']), fecha_hora=data['fecha_hora'], zona=data['zona'], tipo=data['tipo'], descripcion=data['descripcion'], personal_involucrado=data['personal_involucrado'], respuesta_implementada=data['respuesta_implementada'], tiempo_reaccion=int(data['tiempo_reaccion'] or 0))
            messagebox.showinfo('Éxito', f'{cnt} registro(s) actualizados.')
            self.load_incidentes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo actualizar: {e}')

    def delete_incidente(self):
        try:
            if not hasattr(self, 'selected_incidente_id'):
                messagebox.showwarning('Seleccionar','Seleccione primero un incidente de la lista.')
                return
            if messagebox.askyesno('Confirmar','¿Eliminar incidente seleccionado?'):
                cnt = delete_incidente(self.selected_incidente_id)
                messagebox.showinfo('Éxito', f'{cnt} registro(s) eliminados.')
                self.load_incidentes()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo eliminar: {e}')

    def export_incidentes_excel(self):
        try:
            rows = read_incidentes(1000)
            if not rows:
                messagebox.showwarning("Atención", "No hay datos para exportar.")
                return

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar como...",
                initialfile="incidentes.xlsx"
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Incidentes"

            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Exportado a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")

    # ---------------- Instalaciones view ----------------
    def show_instalaciones(self):
        self.clear_content()
        frame = tk.Frame(self.content, bg='#ecf0f1', padx=10, pady=10)
        frame.pack(fill='both', expand=True)

        # Formulario crear instalación
        form = tk.LabelFrame(frame, text='Crear / Actualizar Instalación', padx=10, pady=10)
        form.pack(fill='x', padx=10, pady=10)
        entries = {}
        fields = [('cliente_id', 'Cliente ID'), ('tipo', 'Tipo'), ('direccion', 'Dirección'),
                  ('coordenadas', 'Coordenadas'), ('nivel_seguridad', 'Nivel Seguridad')]
        for i, (k, label) in enumerate(fields):
            tk.Label(form, text=label).grid(row=i, column=0, sticky='w')
            ent = tk.Entry(form, width=50)
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[k] = ent

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=8)
        tk.Button(btn_frame, text='Crear Instalación', command=lambda: self.create_instalacion(entries)).pack(
            side='left', padx=5)
        tk.Button(btn_frame, text='Actualizar Instalación', command=lambda: self.update_instalacion(entries)).pack(
            side='left', padx=5)
        tk.Button(btn_frame, text='Eliminar Instalación', command=self.delete_instalacion).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Refrescar Lista', command=self.load_instalaciones).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Exportar Excel', command=self.export_instalaciones_excel).pack(side='left', padx=5)

        # Tabla de instalaciones
        table_frame = tk.LabelFrame(frame, text='Instalaciones - Lista', padx=5, pady=5)
        table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        cols = ('instalacion_id', 'cliente_id', 'tipo', 'direccion', 'coordenadas', 'nivel_seguridad')
        tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120)
        tree.pack(fill='both', expand=True)
        tree.bind('<<TreeviewSelect>>', lambda e: self.on_instalacion_select(e, tree, entries))

        self.inst_tree = tree
        self.load_instalaciones()

    def create_instalacion(self, entries):
        try:
            data = {k: entries[k].get().strip() for k in entries}
            if not data['cliente_id']:
                messagebox.showwarning('Validación', 'Cliente ID es requerido.')
                return
            nid = create_instalacion(data['cliente_id'], data['tipo'], data['direccion'], data['coordenadas'],
                                     data['nivel_seguridad'])
            messagebox.showinfo('Éxito', f'Instalación creada con ID {nid}')
            self.load_instalaciones()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo crear instalación: {e}')

    def load_instalaciones(self):
        try:
            rows = read_instalaciones(200)
            self.inst_tree.delete(*self.inst_tree.get_children())
            for r in rows:
                self.inst_tree.insert('', 'end', values=(
                r['instalacion_id'], r['cliente_id'], r['tipo'], r['direccion'], r.get('coordenadas', ''),
                r.get('nivel_seguridad', '')))
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo cargar instalaciones: {e}')

    def on_instalacion_select(self, event, tree, entries):
        try:
            sel = tree.selection()
            if not sel: return
            vals = tree.item(sel[0], 'values')
            # map to entries
            entries['cliente_id'].delete(0, 'end')
            entries['cliente_id'].insert(0, vals[1])
            entries['tipo'].delete(0, 'end')
            entries['tipo'].insert(0, vals[2])
            entries['direccion'].delete(0, 'end')
            entries['direccion'].insert(0, vals[3])
            entries['coordenadas'].delete(0, 'end')
            entries['coordenadas'].insert(0, vals[4])
            entries['nivel_seguridad'].delete(0, 'end')
            entries['nivel_seguridad'].insert(0, vals[5])
            # store selected id
            self.selected_instalacion_id = vals[0]
        except Exception as e:
            print('on_instalacion_select error', e)

    def update_instalacion(self, entries):
        try:
            if not hasattr(self, 'selected_instalacion_id'):
                messagebox.showwarning('Seleccionar', 'Seleccione primero una instalación de la lista.')
                return
            data = {k: entries[k].get().strip() for k in entries}
            cnt = update_instalacion(self.selected_instalacion_id, cliente_id=data['cliente_id'], tipo=data['tipo'],
                                     direccion=data['direccion'], coordenadas=data['coordenadas'],
                                     nivel_seguridad=data['nivel_seguridad'])
            messagebox.showinfo('Éxito', f'{cnt} registro(s) actualizados.')
            self.load_instalaciones()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo actualizar: {e}')

    def delete_instalacion(self):
        try:
            if not hasattr(self, 'selected_instalacion_id'):
                messagebox.showwarning('Seleccionar', 'Seleccione primero una instalación de la lista.')
                return
            if messagebox.askyesno('Confirmar', '¿Eliminar instalación seleccionada?'):
                cnt = delete_instalacion(self.selected_instalacion_id)
                messagebox.showinfo('Éxito', f'{cnt} registro(s) eliminados.')
                self.load_instalaciones()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo eliminar: {e}')
    def export_instalaciones_excel(self):
        try:
            rows = read_instalaciones(1000)
            if not rows:
                messagebox.showwarning("Atención", "No hay datos para exportar.")
                return

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar como...",
                initialfile="instalaciones.xlsx"
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Instalaciones"

            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Exportado a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")


    # ---------------- Personal view ----------------
    def show_personal_view(self):
        self.clear_content()
        frame = tk.Frame(self.content, bg='#ecf0f1', padx=10, pady=10)
        frame.pack(fill='both', expand=True)

        form = tk.LabelFrame(frame, text='Registrar Personal Seguridad', padx=10, pady=10)
        form.pack(fill='x', padx=10, pady=10)
        entries = {}
        fields = [('nombres', 'Nombres'), ('apellidos', 'Apellidos'), ('documento_identidad', 'Documento Identidad'),
                  ('telefono', 'Teléfono'), ('formacion', 'Formación'),
                  ('evaluacion_periodica', 'Evaluación Periódica')]
        for i, (k, label) in enumerate(fields):
            tk.Label(form, text=label).grid(row=i, column=0, sticky='w')
            ent = tk.Entry(form, width=50)
            ent.grid(row=i, column=1, padx=5, pady=2)
            entries[k] = ent

        btn_frame = tk.Frame(form)
        btn_frame.grid(row=len(fields), column=0, columnspan=2, pady=8)
        tk.Button(btn_frame, text='Crear Personal', command=lambda: self.create_personal(entries)).pack(side='left',
                                                                                                        padx=5)
        tk.Button(btn_frame, text='Actualizar Personal', command=lambda: self.update_personal(entries)).pack(
            side='left', padx=5)
        tk.Button(btn_frame, text='Eliminar Personal', command=self.delete_personal).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Refrescar Lista', command=self.load_personal).pack(side='left', padx=5)
        tk.Button(btn_frame, text='Exportar Excel', command=self.export_personal_excel).pack(side='left', padx=5)

        table_frame = tk.LabelFrame(frame, text='Personal Seguridad - Lista', padx=5, pady=5)
        table_frame.pack(fill='both', expand=True, padx=10, pady=10)
        cols = (
        'empleado_id', 'nombres', 'apellidos', 'documento_identidad', 'telefono', 'formacion', 'evaluacion_periodica')
        tree = ttk.Treeview(table_frame, columns=cols, show='headings')
        for c in cols:
            tree.heading(c, text=c)
            tree.column(c, width=120)
        tree.pack(fill='both', expand=True)
        tree.bind('<<TreeviewSelect>>', lambda e: self.on_personal_select(e, tree, entries))

        self.personal_tree = tree
        self.load_personal()

    def create_personal(self, entries):
        try:
            data = {k: entries[k].get().strip() for k in entries}
            # Validaciones básicas
            if not data['nombres'] or not data['apellidos'] or not data['documento_identidad']:
                messagebox.showwarning('Validación', 'Nombres, apellidos y documento son requeridos.')
                return
            nid = create_personal(
                data['nombres'], data['apellidos'], data['documento_identidad'],
                data['telefono'], data['formacion'], data['evaluacion_periodica']
            )
            messagebox.showinfo('Éxito', f'Personal creado con ID {nid}')
            self.load_personal()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo crear personal: {e}')

    def load_personal(self):
        try:
            rows = read_personal(200)
            self.personal_tree.delete(*self.personal_tree.get_children())
            for r in rows:
                self.personal_tree.insert('', 'end', values=(
                    r['empleado_id'], r['nombres'], r['apellidos'], r['documento_identidad'],
                    r.get('telefono', ''), r.get('formacion', ''), r.get('evaluacion_periodica', '')
                ))
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo cargar personal: {e}')

    def on_personal_select(self, event, tree, entries):
        try:
            sel = tree.selection()
            if not sel: return
            vals = tree.item(sel[0], 'values')
            # Mapear valores a los campos del formulario
            entries['nombres'].delete(0, 'end')
            entries['nombres'].insert(0, vals[1])
            entries['apellidos'].delete(0, 'end')
            entries['apellidos'].insert(0, vals[2])
            entries['documento_identidad'].delete(0, 'end')
            entries['documento_identidad'].insert(0, vals[3])
            entries['telefono'].delete(0, 'end')
            entries['telefono'].insert(0, vals[4])
            entries['formacion'].delete(0, 'end')
            entries['formacion'].insert(0, vals[5])
            entries['evaluacion_periodica'].delete(0, 'end')
            entries['evaluacion_periodica'].insert(0, vals[6])
            # Guardar ID seleccionado
            self.selected_personal_id = vals[0]
        except Exception as e:
            print('on_personal_select error', e)

    def update_personal(self, entries):
        try:
            if not hasattr(self, 'selected_personal_id'):
                messagebox.showwarning('Seleccionar', 'Seleccione primero un personal de la lista.')
                return
            data = {k: entries[k].get().strip() for k in entries}
            cnt = update_personal(
                self.selected_personal_id,
                nombres=data['nombres'],
                apellidos=data['apellidos'],
                documento_identidad=data['documento_identidad'],
                telefono=data['telefono'],
                formacion=data['formacion'],
                evaluacion_periodica=data['evaluacion_periodica']
            )
            messagebox.showinfo('Éxito', f'{cnt} registro(s) actualizados.')
            self.load_personal()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo actualizar: {e}')

    def delete_personal(self):
        try:
            if not hasattr(self, 'selected_personal_id'):
                messagebox.showwarning('Seleccionar', 'Seleccione primero un personal de la lista.')
                return
            if messagebox.askyesno('Confirmar', '¿Eliminar personal seleccionado?'):
                cnt = delete_personal(self.selected_personal_id)
                messagebox.showinfo('Éxito', f'{cnt} registro(s) eliminados.')
                self.load_personal()
        except Exception as e:
            messagebox.showerror('Error', f'No se pudo eliminar: {e}')
    def export_personal_excel(self):
        try:
            rows = read_personal(1000)
            if not rows:
                messagebox.showwarning("Atención", "No hay datos para exportar.")
                return

            filepath = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel", "*.xlsx")],
                title="Guardar como...",
                initialfile="personal.xlsx"
            )
            if not filepath:
                return

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Personal"

            ws.append(list(rows[0].keys()))
            for r in rows:
                ws.append(list(r.values()))

            wb.save(filepath)
            messagebox.showinfo("Éxito", f"Exportado a {filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo exportar: {e}")


if __name__ == '__main__':
    app = SecureSysApp()
    app.mainloop()
